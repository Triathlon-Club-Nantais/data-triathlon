"""
Extract provider URLs from the TCN registration xlsx, grouped by provider.
Outputs a JSON dict ready to append to tests/e2e/fixtures/providers.json.
"""
import sys, json, re
from pathlib import Path
from urllib.parse import urlparse, parse_qs

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

# Locate xlsx
DATA_DIR = Path(r"C:\dev\data-triathlon\.claude\data")
xlsx_files = list(DATA_DIR.glob("*.xlsx"))
if not xlsx_files:
    sys.exit(f"No xlsx found in {DATA_DIR}")
xlsx_path = xlsx_files[0]
print(f"[info] Loading: {xlsx_path.name}")

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
print(f"[info] Columns: {headers}")
print(f"[info] Rows: {ws.max_row - 1}")

url_pat = re.compile(r"https?://[^\s\]\"'<>|]+", re.I)

# Collect all URLs with their row context
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Look for URL cell and name cell
    row_str = {headers[i]: str(v).strip() if v else "" for i, v in enumerate(row)}
    row_urls = []
    row_names = []

    for key, val in row_str.items():
        if not val:
            continue
        # Collect URLs
        found = url_pat.findall(val)
        row_urls.extend(found)
        # Collect name hints from "Nom" / "Prénom" / "Nom complet" type columns
        key_l = key.lower()
        if any(x in key_l for x in ("nom", "prenom", "prénom", "athlete", "athlète", "participant")):
            row_names.append(val)

    if not row_urls:
        continue

    # Best name guess: first non-URL value that looks like a name
    name_hint = " ".join(row_names[:2]).strip() if row_names else ""
    for url in row_urls:
        records.append({"url": url, "name": name_hint, "row": row_str})

print(f"\n[info] Total URL records: {len(records)}")

def classify_provider(url: str) -> str:
    u = url.lower()
    if "klikego.com" in u:
        return "klikego"
    if "live.breizhchrono.com" in u:
        return "breizhchrono_live"
    if "breizhchrono.com/detail-de-la-course" in u:
        return "breizhchrono_dead"
    if "breizhchrono.com" in u:
        return "breizhchrono"
    if "chronosmetron.com" in u or "wiclax-results.com" in u or ("wiclax" in u and "g-live" in u):
        return "wiclax"
    if "timepulse.fr" in u:
        return "timepulse"
    if "prolivesport.fr" in u and "eventid" in u:
        return "prolivesport"
    if "sportinnovation.fr" in u:
        return "sportinnovation"
    return "unknown"

def extract_search(url: str, name_hint: str) -> str:
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    # Klikego/breizhchrono use ?search=
    s = params.get("search", [""])[0].strip()
    if s:
        return s
    # Timepulse uses ?search= or bib
    # Wiclax uses ?B= (bib) or search
    b = params.get("b", params.get("B", [""]))[0].strip()
    if not s and not b and name_hint:
        # Extract surname from name hint (first uppercase word)
        parts = name_hint.upper().split()
        return parts[0] if parts else ""
    return s or ""

grouped: dict[str, list] = {}
for rec in records:
    url = rec["url"].rstrip("/")
    # Remove trailing garbage sometimes in xlsx
    url = re.sub(r"[,;]+$", "", url)
    provider = classify_provider(url)
    if provider == "unknown":
        continue
    search = extract_search(url, rec["name"])
    key = (provider, url, search)
    if key not in {(p, e["url"], e["search"]) for p, entries in grouped.items() for e in entries}:
        grouped.setdefault(provider, []).append({
            "provider": provider,
            "url": url,
            "search": search,
            "name_hint": rec["name"],
        })

print("\n[info] Provider counts from xlsx:")
for p, entries in sorted(grouped.items()):
    print(f"  {p}: {len(entries)}")

# Save as JSON for inspection
out_path = Path(r"C:\dev\data-triathlon\backend\xlsx_urls.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(grouped, f, ensure_ascii=False, indent=2)
print(f"\n[info] Saved to {out_path}")
