"""Source Google Sheet : téléchargement du CSV, extraction et normalisation des liens.

Aucun accès DB, aucun scraping — juste la lecture de la source d'entrée de
l'import de masse.
"""
import csv
import io
from dataclasses import dataclass
from urllib.parse import urlparse, urlunparse

from app.core import http
from app.core.exceptions import DomainError
from app.scrapers import registry

DEFAULT_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1rtiVRFOQUGcaWCTDPTR4xA9UL22UsWosKjsYMcRMsew/export?format=csv&gid=1961918487"
)
LINK_HEADER = "Donne-nous un lien pour accéder aux résultats."
LINK_COLUMN_FALLBACK_INDEX = 9  # 10e colonne, repli si l'en-tête n'est pas trouvé


def normalize_url(url: str) -> str:
    """Normalise pour la déduplication : trim, host en minuscule, slash final et
    fragment supprimés. La query est conservée (elle distingue deux heats)."""
    parsed = urlparse(url.strip())
    scheme = parsed.scheme.lower()
    netloc = parsed.netloc.lower()
    path = parsed.path.rstrip("/") or "/"
    return urlunparse((scheme, netloc, path, parsed.params, parsed.query, ""))


def dedupe_links(links: list[str]) -> list[str]:
    """Dédoublonne par URL normalisée en conservant l'ordre et la forme d'origine.

    `setdefault` garde la **première** forme rencontrée, pas la dernière : c'est
    ce qui interdit un simple `dict.fromkeys`, qui garderait la clé normalisée.
    Même motif que `rescrape_service._dedupe_par_url`.
    """
    uniques: dict[str, str] = {}
    for url in links:
        uniques.setdefault(normalize_url(url), url)
    return list(uniques.values())


class UnsupportedFileError(DomainError):
    status_code = 422
    message = "Format non pris en charge. Seuls les fichiers .csv et .xlsx sont lus."


class UnreadableFileError(DomainError):
    status_code = 422
    message = "Fichier illisible. Vérifiez qu'il s'agit bien d'un .csv ou d'un .xlsx."


def read_table(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """(en-têtes, lignes) d'un `.csv` ou d'un `.xlsx`, **en mémoire**.

    Le fichier téléversé n'est jamais écrit côté serveur (FR-011) : il arrive en
    octets et repart en listes. Le navigateur, lui, le garde entre l'appel qui
    liste les colonnes et celui qui lance le batch.

    Toute cellule ressort en `str` : `openpyxl` rend `None` pour une case vide
    et un `int` pour un nombre, et les deux feraient tomber `.startswith` en
    aval. Un en-tête vide devient « Colonne N » — une colonne qu'on ne peut pas
    nommer est une colonne qu'on ne peut pas désigner.
    """
    suffixe = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffixe == "csv":
        lignes = _lignes_csv(content)
    elif suffixe == "xlsx":
        lignes = _lignes_xlsx(content)
    else:
        raise UnsupportedFileError

    if not lignes:
        return [], []
    # Rang **à partir de 1** : c'est un libellé lu par un humain, et « Colonne 0 »
    # ne désigne rien pour lui. L'`index` du contrat, lui, reste à partir de 0 —
    # il ne s'affiche jamais.
    entetes = [
        cellule or f"Colonne {rang}" for rang, cellule in enumerate(lignes[0], 1)
    ]
    return entetes, lignes[1:]


def _lignes_csv(content: bytes) -> list[list[str]]:
    # Les exports d'un tableur français arrivent souvent en cp1252 ; échouer sur
    # un accent rendrait l'écran inutilisable un jour sur deux. `cp1252` accepte
    # tout octet, donc ce repli ne lève jamais.
    try:
        texte = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        texte = content.decode("cp1252")
    return [
        [(cellule or "").strip() for cellule in ligne]
        for ligne in csv.reader(io.StringIO(texte))
    ]


def _lignes_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        # `read_only` ne charge pas la feuille entière en mémoire ; `data_only`
        # rend la valeur calculée d'une formule plutôt que la formule elle-même
        # — un lien produit par `=CONCATENER(...)` sortirait sinon en texte de
        # formule, et ne compterait pour aucun lien.
        classeur = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as erreur:
        raise UnreadableFileError from erreur
    try:
        return [
            ["" if cellule is None else str(cellule).strip() for cellule in ligne]
            for ligne in classeur.active.iter_rows(values_only=True)
        ]
    finally:
        classeur.close()


@dataclass(frozen=True)
class ColumnLinks:
    """Ce qu'une colonne porte, vu avant tout lancement."""

    links: list[str]           # dédoublonnés, ordre d'origine conservé
    supported: list[str]       # ceux qu'un scraper du registre reconnaît
    ignored_by_host: dict[str, int]  # les autres, groupés par hôte
    rows_without_link: int     # lignes non vides ne portant pas d'URL

    @property
    def count(self) -> int:
        """Le compte affiché en face de la colonne dans le sélecteur."""
        return len(self.links)


def links_in_column(rows: list[list[str]], index: int) -> ColumnLinks:
    """Les liens d'une colonne, dédoublonnés et partagés supportés / ignorés.

    Le partage est fait **avant** le lancement, et c'est ce qui permet à l'écran
    d'annoncer ce qui ne sera jamais soumis. Un lien ignoré n'est ni un succès
    ni un échec ; le confondre avec l'un des deux fausse tous les compteurs.
    """
    bruts: list[str] = []
    sans_lien = 0
    for row in rows:
        valeur = row[index].strip() if index < len(row) else ""
        if valeur.startswith("http"):
            bruts.append(valeur)
        elif valeur:
            sans_lien += 1

    links = dedupe_links(bruts)
    supported = [url for url in links if registry.is_supported(url)]
    ignores: dict[str, int] = {}
    for url in links:
        if url not in supported:
            ignores[host_of(url)] = ignores.get(host_of(url), 0) + 1
    return ColumnLinks(
        links=links,
        supported=supported,
        ignored_by_host=ignores,
        rows_without_link=sans_lien,
    )


def parse_sheet_csv(csv_text: str) -> tuple[list[str], int]:
    """Extrait la colonne des liens du CSV — **défauts de la commande CLI**.

    `LINK_HEADER` et l'index 9 sont les repères du Sheet historique du club, et
    ils restent ici : la CLI, elle, n'a personne pour désigner une colonne.
    L'écran d'administration passe par `read_table` + `links_in_column` et
    demande la colonne, ce qui est précisément la différence entre les deux
    guichets.

    Ne dédoublonne pas : `bulk_import_service` le fait plus loin, et le compte
    brut de liens fait partie de son bilan.
    """
    entetes, lignes = read_table(csv_text.encode("utf-8"), "sheet.csv")
    if not entetes:
        return [], 0
    try:
        col = entetes.index(LINK_HEADER)
    except ValueError:
        col = LINK_COLUMN_FALLBACK_INDEX

    links: list[str] = []
    sans_lien = 0
    for row in lignes:
        value = row[col].strip() if col < len(row) else ""
        if value.startswith("http"):
            links.append(value)
        elif any(cell.strip() for cell in row):
            sans_lien += 1
    return links, sans_lien


def host_of(url: str) -> str:
    """Host en minuscule, pour grouper les liens ignorés dans le rapport."""
    return (urlparse(url).netloc or "").lower() or "(inconnu)"


def download_csv(url: str) -> str:
    """Télécharge le CSV public du Sheet (sans auth).

    L'export d'un Google Sheet redirige vers `googleusercontent.com`, un autre
    domaine : c'est le cas qui a fait écarter l'allowlist de hosts par provider
    au profit de la politique par classe d'IP (#101).
    """
    with http.client(timeout=30) as client:
        resp = client.get(url)
        resp.raise_for_status()
        return resp.text
